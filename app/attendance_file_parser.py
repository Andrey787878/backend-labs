from __future__ import annotations

from datetime import date, datetime, time
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.dto import AttendanceRowDTO


class AttendanceParseError(ValueError):
    """Ошибка чтения или нормализации Excel-файла посещаемости."""


class AttendanceFileParser:
    """Читает лист 'Посещаемость' и преобразует строки Excel в DTO."""

    SHEET_NAME = "Посещаемость"
    REQUIRED_HEADERS: dict[str, tuple[str, ...]] = {
        "group_name": ("группа",),
        "subgroup": ("подгруппа",),
        "student_name": ("фио",),
        "date": ("дата",),
        "time": ("время",),
        "lesson_type": ("тип занятия", "тип"),
        "lesson_number": ("номер занятия", "номер"),
        "visited": ("посещение",),
        "success_labs": ("выполнено лаб",),
        "already_auto_credit": ("зачет автоматом",),
    }

    def parse(self, file_content: bytes) -> list[AttendanceRowDTO]:
        """Возвращает нормализованные строки из .xlsx-файла."""
        try:
            workbook = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        except (InvalidFileException, OSError, ValueError) as exc:
            raise AttendanceParseError("Не удалось прочитать .xlsx файл.") from exc

        if self.SHEET_NAME not in workbook.sheetnames:
            raise AttendanceParseError("В файле должен быть лист 'Посещаемость'.")

        worksheet = workbook[self.SHEET_NAME]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            raise AttendanceParseError("Лист 'Посещаемость' пустой.")

        header_map = self._build_header_map(header_row)
        rows: list[AttendanceRowDTO] = []

        for row_number, raw_row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if self._is_empty_row(raw_row):
                continue
            rows.append(self._parse_row(row_number=row_number, raw_row=raw_row, header_map=header_map))

        return rows

    def _build_header_map(self, header_row: tuple[Any, ...]) -> dict[str, int]:
        normalized_headers = {
            self._normalize_header(value): index
            for index, value in enumerate(header_row)
            if value is not None and str(value).strip()
        }

        missing_headers = []
        header_map: dict[str, int] = {}
        for target_field, source_headers in self.REQUIRED_HEADERS.items():
            matched_source_header = next(
                (
                    source_header
                    for source_header in source_headers
                    if source_header in normalized_headers
                ),
                None,
            )
            if matched_source_header is None:
                missing_headers.append(source_headers[0])
                continue
            header_map[target_field] = normalized_headers[matched_source_header]

        if missing_headers:
            missing = ", ".join(missing_headers)
            raise AttendanceParseError(f"В файле отсутствуют обязательные колонки: {missing}.")

        return header_map

    def _parse_row(
        self,
        *,
        row_number: int,
        raw_row: tuple[Any, ...],
        header_map: dict[str, int],
    ) -> AttendanceRowDTO:
        try:
            group_name = self._required_text(self._value(raw_row, header_map["group_name"]), "Группа")
            student_name = self._required_text(self._value(raw_row, header_map["student_name"]), "ФИО")
            lesson_type = self._normalize_lesson_type(self._value(raw_row, header_map["lesson_type"]))
            return AttendanceRowDTO(
                group_name=group_name,
                subgroup=self._parse_subgroup(self._value(raw_row, header_map["subgroup"])),
                student_name=student_name,
                date=self._parse_date(self._value(raw_row, header_map["date"])),
                time=self._parse_time(self._value(raw_row, header_map["time"])),
                lesson_type=lesson_type,
                lesson_number=self._parse_positive_int(
                    self._value(raw_row, header_map["lesson_number"]),
                    "Номер занятия",
                ),
                visited=self._parse_bool(self._value(raw_row, header_map["visited"]), "Посещение"),
                success_labs=self._parse_non_negative_int(
                    self._value(raw_row, header_map["success_labs"]),
                    "Выполнено лаб",
                ),
                already_auto_credit=self._parse_bool(
                    self._value(raw_row, header_map["already_auto_credit"]),
                    "Зачёт автоматом",
                ),
            )
        except AttendanceParseError as exc:
            raise AttendanceParseError(f"Строка {row_number}: {exc}") from exc

    @staticmethod
    def _normalize_header(value: Any) -> str:
        return " ".join(str(value).strip().lower().replace("ё", "е").split())

    @staticmethod
    def _is_empty_row(raw_row: tuple[Any, ...]) -> bool:
        return all(value is None or str(value).strip() == "" for value in raw_row)

    @staticmethod
    def _value(raw_row: tuple[Any, ...], index: int) -> Any:
        return raw_row[index] if index < len(raw_row) else None

    @staticmethod
    def _required_text(value: Any, field_name: str) -> str:
        if value is None:
            raise AttendanceParseError(f"{field_name} не должен быть пустым.")
        normalized = str(value).strip()
        if not normalized:
            raise AttendanceParseError(f"{field_name} не должен быть пустым.")
        return normalized

    def _parse_subgroup(self, value: Any) -> int:
        if value is None or str(value).strip() == "":
            return 1
        return self._parse_positive_int(value, "Подгруппа")

    @staticmethod
    def _parse_date(value: Any) -> str:
        if isinstance(value, datetime):
            return value.strftime("%d.%m.%Y")
        if isinstance(value, date):
            return value.strftime("%d.%m.%Y")
        if value is None:
            raise AttendanceParseError("Дата не должна быть пустой.")

        text = str(value).strip()
        for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).strftime("%d.%m.%Y")
            except ValueError:
                continue
        raise AttendanceParseError("Дата должна быть в формате ДД.ММ.ГГГГ.")

    @staticmethod
    def _parse_time(value: Any) -> str:
        if isinstance(value, datetime):
            return value.strftime("%H:%M")
        if isinstance(value, time):
            return value.strftime("%H:%M")
        if value is None:
            raise AttendanceParseError("Время не должно быть пустым.")

        text = str(value).strip()
        for fmt in ("%H:%M", "%H:%M:%S"):
            try:
                return datetime.strptime(text, fmt).strftime("%H:%M")
            except ValueError:
                continue
        raise AttendanceParseError("Время должно быть в формате ЧЧ:ММ.")

    @staticmethod
    def _normalize_lesson_type(value: Any) -> str:
        if value is None:
            raise AttendanceParseError("Тип занятия не должен быть пустым.")
        normalized = str(value).strip().lower()
        if normalized not in {"lect", "lab"}:
            raise AttendanceParseError("Тип занятия должен быть 'lect' или 'lab'.")
        return normalized

    @classmethod
    def _parse_positive_int(cls, value: Any, field_name: str) -> int:
        parsed = cls._parse_int(value=value, field_name=field_name)
        if parsed < 1:
            raise AttendanceParseError(f"{field_name} должен быть положительным числом.")
        return parsed

    @classmethod
    def _parse_non_negative_int(cls, value: Any, field_name: str) -> int:
        parsed = cls._parse_int(value=value, field_name=field_name)
        if parsed < 0:
            raise AttendanceParseError(f"{field_name} не может быть отрицательным.")
        return parsed

    @staticmethod
    def _parse_int(value: Any, field_name: str) -> int:
        if value is None or str(value).strip() == "":
            raise AttendanceParseError(f"{field_name} не должен быть пустым.")
        try:
            parsed_float = float(str(value).strip())
        except ValueError as exc:
            raise AttendanceParseError(f"{field_name} должен быть целым числом.") from exc

        if not parsed_float.is_integer():
            raise AttendanceParseError(f"{field_name} должен быть целым числом.")
        return int(parsed_float)

    @staticmethod
    def _parse_bool(value: Any, field_name: str) -> bool:
        if isinstance(value, bool):
            return value
        if value is None or str(value).strip() == "":
            raise AttendanceParseError(f"{field_name} не должен быть пустым.")

        normalized = str(value).strip().lower().replace("ё", "е")
        if normalized in {"1", "да", "yes", "true", "y", "истина"}:
            return True
        if normalized in {"0", "нет", "no", "false", "n", "ложь"}:
            return False
        raise AttendanceParseError(f"{field_name} должен быть 1/0 или да/нет.")
